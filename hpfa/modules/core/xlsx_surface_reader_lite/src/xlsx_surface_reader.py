from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

MODULE_ID = "xlsx_surface_reader_lite_v1"
CANONICAL_EVENT_COUNT = "UNKNOWN"
CLAIM_CEILING = "XLSX_SURFACE_AUDIT_ONLY"
MAX_SURFACE_CELLS = 2_000_000
OUT = {
    "main": "xlsx_surface_audit_lite_v1.json",
    "summary": "xlsx_surface_audit_lite_v1.txt",
    "analyst": "xlsx_surface_analyst_audit_lite_v1.txt",
}

IDENTITY_ALIASES = {
    "player": {
        "player",
        "player_name",
        "name",
        "oyuncu",
        "oyuncu_adi",
        "oyuncu_adı",
    },
    "team": {
        "team",
        "team_name",
        "club",
        "takim",
        "takım",
        "takim_adi",
        "takım_adı",
    },
    "position": {"position", "pos", "role", "pozisyon", "mevki"},
    "minutes": {
        "minutes",
        "minutes_played",
        "mins",
        "minute",
        "dakika",
        "oynanan_dakika",
    },
    "shirt_number": {
        "no",
        "number",
        "shirt_number",
        "jersey_number",
        "forma_numarasi",
        "forma_numarası",
    },
}


def norm(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^\w]+", "_", text, flags=re.UNICODE)
    return re.sub(r"_+", "_", text).strip("_")


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def jsonable(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, float) and not math.isfinite(value):
        return str(value)
    return value


def stable_key(value: Any) -> str:
    return json.dumps(jsonable(value), ensure_ascii=False, sort_keys=True, default=str)


def value_kind(value: Any) -> str:
    if is_blank(value):
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return "date_or_time"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    if isinstance(value, str):
        return "string"
    return type(value).__name__.casefold()


def inferred_type(values: Iterable[Any]) -> str:
    kinds = {value_kind(value) for value in values if not is_blank(value)}
    if not kinds:
        return "unknown"
    if len(kinds) == 1:
        return next(iter(kinds))
    if kinds <= {"number", "boolean"}:
        return "number_or_boolean"
    return "mixed"


def header_index(rows: list[list[Any]], scan_limit: int = 50) -> int | None:
    best: tuple[int, int] | None = None
    for index, row in enumerate(rows[:scan_limit]):
        present = [str(value).strip() for value in row if not is_blank(value)]
        if len(present) < 2:
            continue
        unique = len({norm(value) for value in present if norm(value)})
        textual = sum(
            1
            for value in present
            if not re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", value)
        )
        if unique < 2 or textual < 1:
            continue
        score = len(present) * 5 + unique * 2 + textual
        candidate = (score, -index)
        if best is None or candidate > best:
            best = candidate
    return -best[1] if best is not None else None


def identity_role(normalized_header: str) -> str | None:
    for role, aliases in IDENTITY_ALIASES.items():
        if normalized_header in aliases:
            return role
    return None


def sheet_rows(ws_values: Any, ws_formula: Any) -> tuple[list[list[Any]], int, int]:
    max_row = max(int(ws_values.max_row or 0), int(ws_formula.max_row or 0))
    max_col = max(int(ws_values.max_column or 0), int(ws_formula.max_column or 0))
    if max_row * max_col > MAX_SURFACE_CELLS:
        raise ValueError("xlsx_surface_cell_budget_exceeded")
    rows: list[list[Any]] = []
    actual_width = 0
    last_nonblank_row = 0
    for row_number in range(1, max_row + 1):
        row: list[Any] = []
        row_has_content = False
        for column_number in range(1, max_col + 1):
            value_cell = ws_values.cell(row=row_number, column=column_number)
            formula_cell = ws_formula.cell(row=row_number, column=column_number)
            cached = value_cell.value
            formula = formula_cell.value if formula_cell.data_type == "f" else None
            row.append(cached)
            if not is_blank(cached) or formula is not None:
                row_has_content = True
                actual_width = max(actual_width, column_number)
        rows.append(row)
        if row_has_content:
            last_nonblank_row = row_number
    if last_nonblank_row == 0 or actual_width == 0:
        return [], 0, 0
    compact = [row[:actual_width] for row in rows[:last_nonblank_row]]
    return compact, last_nonblank_row, actual_width


def column_profile(
    raw_header: Any,
    normalized_header: str,
    values: list[Any],
    formula_cells: list[Any],
    cached_formula_cells: list[Any],
    number_formats: list[str],
    duplicate_header: bool,
    unnamed_header: bool,
) -> dict[str, Any]:
    present = [value for value in values if not is_blank(value)]
    counts = Counter(value_kind(value) for value in values)
    numeric = [
        float(value)
        for value in present
        if isinstance(value, (int, float))
        and not isinstance(value, bool)
        and math.isfinite(float(value))
    ]
    role = identity_role(normalized_header)
    formulas = [cell for cell in formula_cells if cell is not None]
    missing_formula_cache = sum(
        1
        for formula, cached in zip(formula_cells, cached_formula_cells)
        if formula is not None and is_blank(cached)
    )
    warnings: list[str] = []
    if duplicate_header:
        warnings.append("duplicate_column_name")
    if unnamed_header:
        warnings.append("unnamed_column")
    if missing_formula_cache:
        warnings.append("formula_without_cached_value")
    return {
        "raw_column": "" if raw_header is None else str(raw_header),
        "normalized_column": normalized_header,
        "identity_role_candidate": role,
        "metric_surface_candidate": role is None,
        "inferred_type": inferred_type(values),
        "value_kind_counts": dict(sorted(counts.items())),
        "null_count": counts.get("blank", 0),
        "null_ratio": counts.get("blank", 0) / len(values) if values else 0.0,
        "unique_count": len({stable_key(value) for value in present}),
        "minimum": min(numeric) if numeric else None,
        "maximum": max(numeric) if numeric else None,
        "formula_cell_count": len(formulas),
        "formula_without_cached_value_count": missing_formula_cache,
        "percent_number_format_count": sum(
            1 for fmt in number_formats if "%" in str(fmt or "")
        ),
        "percent_header_candidate": (
            "%" in str(raw_header or "")
            or "percent" in normalized_header
            or normalized_header.endswith("_pct")
        ),
        "claim_ceiling": "AGGREGATE_FIELD_SURFACE_ONLY",
        "parse_warning": warnings,
    }


def _examples(values: list[Any], limit: int = 10) -> list[Any]:
    result: list[Any] = []
    seen: set[str] = set()
    for value in values:
        if is_blank(value):
            continue
        key = stable_key(value)
        if key in seen:
            continue
        seen.add(key)
        result.append(jsonable(value))
        if len(result) >= limit:
            break
    return result


def inspect_sheet(ws_formula: Any, ws_values: Any, source_role: str) -> dict[str, Any]:
    state = str(getattr(ws_formula, "sheet_state", "visible") or "visible")
    base = {
        "sheet_name": ws_formula.title,
        "sheet_state": state,
        "source_role": source_role,
        "claim_ceiling": CLAIM_CEILING,
        "canonical_event_count": CANONICAL_EVENT_COUNT,
        "production_release": False,
    }
    try:
        rows, actual_row_count, actual_column_count = sheet_rows(
            ws_values, ws_formula
        )
    except ValueError as exc:
        return base | {
            "status": "FAIL_CLOSED",
            "hard_block_hits": [str(exc)],
            "parse_warnings": [],
            "surface_row_count": 0,
            "visible_column_count": 0,
        }
    if not rows:
        return base | {
            "status": "REVIEW_REQUIRED" if state != "visible" else "FAIL_CLOSED",
            "hard_block_hits": [] if state != "visible" else ["visible_sheet_empty"],
            "parse_warnings": ["hidden_sheet_empty"] if state != "visible" else [],
            "surface_row_count": 0,
            "visible_column_count": 0,
        }

    selected_header = header_index(rows)
    if selected_header is None:
        return base | {
            "status": "REVIEW_REQUIRED" if state != "visible" else "FAIL_CLOSED",
            "hard_block_hits": [] if state != "visible" else ["header_not_found"],
            "parse_warnings": (
                ["hidden_sheet_header_not_found"] if state != "visible" else []
            ),
            "surface_row_count": max(0, len(rows) - 1),
            "visible_column_count": actual_column_count,
        }

    raw_headers = [
        "" if value is None else str(value).strip()
        for value in rows[selected_header]
    ]
    normalized_headers = [norm(value) for value in raw_headers]
    duplicate_names = {
        name
        for name, count in Counter(normalized_headers).items()
        if name and count > 1
    }
    unnamed_indices = [index for index, value in enumerate(raw_headers) if not value]
    body = [
        row
        for row in rows[selected_header + 1 :]
        if any(not is_blank(value) for value in row)
    ]

    formula_count = 0
    formula_without_cache_count = 0
    formula_coordinates: list[str] = []
    profiles: list[dict[str, Any]] = []
    identity_values: dict[str, list[Any]] = {
        role: [] for role in IDENTITY_ALIASES
    }
    metric_inventory: list[dict[str, Any]] = []

    for index in range(actual_column_count):
        values = [
            row[index] if index < len(row) else None
            for row in body
        ]
        formulas: list[Any] = []
        cached: list[Any] = []
        formats: list[str] = []
        for row_number in range(selected_header + 2, actual_row_count + 1):
            formula_cell = ws_formula.cell(
                row=row_number, column=index + 1
            )
            cached_cell = ws_values.cell(
                row=row_number, column=index + 1
            )
            formula = (
                formula_cell.value
                if formula_cell.data_type == "f"
                else None
            )
            formulas.append(formula)
            cached.append(cached_cell.value)
            formats.append(
                str(cached_cell.number_format or formula_cell.number_format or "")
            )
            if formula is not None:
                formula_count += 1
                if len(formula_coordinates) < 20:
                    formula_coordinates.append(formula_cell.coordinate)
                if is_blank(cached_cell.value):
                    formula_without_cache_count += 1

        normalized = (
            normalized_headers[index]
            if index < len(normalized_headers)
            else ""
        )
        profile = column_profile(
            raw_headers[index] if index < len(raw_headers) else "",
            normalized,
            values,
            formulas,
            cached,
            formats,
            normalized in duplicate_names,
            index in unnamed_indices,
        )
        profile["example_values"] = _examples(values, 5)
        profiles.append(profile)
        role = profile["identity_role_candidate"]
        if role:
            identity_values[role].extend(values)
        else:
            metric_inventory.append(
                {
                    "raw_metric_label": profile["raw_column"],
                    "normalized_metric_label": normalized,
                    "inferred_type": profile["inferred_type"],
                    "non_null_count": len(values) - profile["null_count"],
                    "formula_cell_count": profile["formula_cell_count"],
                    "claim_ceiling": (
                        "AGGREGATE_METRIC_LABEL_CANDIDATE_ONLY"
                    ),
                }
            )

    exact_duplicate_row_count = sum(
        count - 1
        for count in Counter(
            tuple(stable_key(value) for value in row)
            for row in body
        ).values()
        if count > 1
    )
    merged_ranges = [str(item) for item in ws_formula.merged_cells.ranges]
    hidden_rows = sorted(
        index
        for index, dim in ws_formula.row_dimensions.items()
        if dim.hidden
    )
    hidden_columns = sorted(
        key
        for key, dim in ws_formula.column_dimensions.items()
        if dim.hidden
    )

    warnings: list[str] = []
    if state != "visible":
        warnings.append("hidden_sheet_not_admitted_to_visible_surface")
    if merged_ranges:
        warnings.append("merged_cells_present")
    if hidden_rows or hidden_columns:
        warnings.append("hidden_rows_or_columns_present")
    if duplicate_names:
        warnings.append("duplicate_column_names")
    if unnamed_indices:
        warnings.append("unnamed_columns")
    if formula_count:
        warnings.append("formula_cells_present")
    if formula_without_cache_count:
        warnings.append("formula_without_cached_value")

    hard_blocks: list[str] = []
    if state == "visible" and not body:
        hard_blocks.append("visible_sheet_has_no_data_rows")
    status = (
        "FAIL_CLOSED"
        if hard_blocks
        else ("REVIEW_REQUIRED" if warnings else "PASS")
    )

    return base | {
        "status": status,
        "header_row_index": selected_header + 1,
        "blank_leading_row_count": sum(
            1
            for row in rows[:selected_header]
            if not any(not is_blank(value) for value in row)
        ),
        "raw_columns": raw_headers,
        "normalized_columns": normalized_headers,
        "visible_column_count": actual_column_count,
        "surface_row_count": len(body),
        "profiled_row_count": len(body),
        "duplicate_column_names": sorted(duplicate_names),
        "unnamed_column_indices": unnamed_indices,
        "column_profiles": profiles,
        "metric_inventory": metric_inventory,
        "identity_binding": {
            role: {
                "binding_status": (
                    "CANDIDATE_ONLY" if _examples(values, 1) else "UNRESOLVED"
                ),
                "candidate_count": len(
                    {
                        stable_key(value)
                        for value in values
                        if not is_blank(value)
                    }
                ),
                "example_candidates": _examples(values),
                "validated_identity": False,
            }
            for role, values in identity_values.items()
        },
        "formula_audit": {
            "formula_cell_count": formula_count,
            "formula_without_cached_value_count": formula_without_cache_count,
            "formula_coordinates_sample": formula_coordinates,
            "formula_evaluation_performed": False,
            "cached_values_used_when_available": True,
        },
        "layout_audit": {
            "actual_row_count_including_header_and_leading_rows": actual_row_count,
            "actual_column_count": actual_column_count,
            "merged_range_count": len(merged_ranges),
            "merged_ranges_sample": merged_ranges[:20],
            "hidden_row_count": len(hidden_rows),
            "hidden_column_count": len(hidden_columns),
            "hidden_rows_sample": hidden_rows[:20],
            "hidden_columns_sample": hidden_columns[:20],
        },
        "exact_duplicate_row_count": exact_duplicate_row_count,
        "duplicate_primary_surface_key": "NOT_EVALUATED_IDENTITY_REQUIRED",
        "hard_block_hits": sorted(set(hard_blocks)),
        "parse_warnings": sorted(set(warnings)),
        "does_not_measure": [
            "canonical_event_truth",
            "validated_team_identity",
            "validated_player_identity",
            "aggregate_definition_truth",
            "cross_format_reconciliation_truth",
            "sequence_truth",
            "phase_truth",
            "tactical_truth",
        ],
    }


def inspect_xlsx_file(path: str | Path, source_role: str) -> dict[str, Any]:
    xlsx_path = Path(path)
    base = {
        "file_name": xlsx_path.name,
        "path": str(xlsx_path),
        "source_role": source_role,
        "canonical_event_count": CANONICAL_EVENT_COUNT,
        "production_release": False,
        "claim_ceiling": CLAIM_CEILING,
    }
    if xlsx_path.suffix.casefold() != ".xlsx":
        return base | {
            "status": "FAIL_CLOSED",
            "hard_block_hits": ["non_xlsx_surface_rejected"],
            "parse_warnings": [],
            "sheets": [],
        }
    if not xlsx_path.is_file():
        return base | {
            "status": "FAIL_CLOSED",
            "hard_block_hits": ["xlsx_file_missing"],
            "parse_warnings": [],
            "sheets": [],
        }
    if not zipfile.is_zipfile(xlsx_path):
        return base | {
            "status": "FAIL_CLOSED",
            "hard_block_hits": ["malformed_xlsx_container"],
            "parse_warnings": [],
            "sheets": [],
        }

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        try:
            formula_book = load_workbook(
                xlsx_path,
                read_only=False,
                data_only=False,
                keep_links=False,
            )
            value_book = load_workbook(
                xlsx_path,
                read_only=False,
                data_only=True,
                keep_links=False,
            )
        except InvalidFileException:
            raise
    except ModuleNotFoundError:
        return base | {
            "status": "FAIL_CLOSED",
            "hard_block_hits": ["openpyxl_dependency_missing"],
            "parse_warnings": [],
            "sheets": [],
        }
    except Exception as exc:
        message = str(exc).casefold()
        block = (
            "encrypted_xlsx"
            if "encrypt" in message or "password" in message
            else "malformed_or_unreadable_xlsx"
        )
        return base | {
            "status": "FAIL_CLOSED",
            "hard_block_hits": [block],
            "parse_warnings": [],
            "error_type": type(exc).__name__,
            "sheets": [],
        }

    try:
        sheet_names = list(formula_book.sheetnames)
        sheets = [
            inspect_sheet(formula_book[name], value_book[name], source_role)
            for name in sheet_names
        ]
        hard_blocks = sorted(
            {
                block
                for sheet in sheets
                for block in sheet.get("hard_block_hits", [])
            }
        )
        warnings = sorted(
            {
                warning
                for sheet in sheets
                for warning in sheet.get("parse_warnings", [])
            }
        )
        if not sheets:
            hard_blocks.append("workbook_has_no_sheets")
        if hard_blocks:
            status = "FAIL_CLOSED"
        elif any(sheet.get("status") == "REVIEW_REQUIRED" for sheet in sheets):
            status = "REVIEW_REQUIRED"
        else:
            status = "PASS"
        return base | {
            "status": status,
            "sheet_count": len(sheets),
            "visible_sheet_count": sum(
                sheet.get("sheet_state") == "visible" for sheet in sheets
            ),
            "hidden_sheet_count": sum(
                sheet.get("sheet_state") != "visible" for sheet in sheets
            ),
            "sheet_names": sheet_names,
            "sheets": sheets,
            "hard_block_hits": hard_blocks,
            "parse_warnings": warnings,
            "workbook_properties": {
                "epoch": str(getattr(formula_book, "epoch", "UNKNOWN")),
                "defined_name_count": len(formula_book.defined_names),
                "calculation_mode": getattr(
                    getattr(formula_book, "calculation", None),
                    "calcMode",
                    None,
                ),
                "external_link_count": len(
                    getattr(formula_book, "_external_links", []) or []
                ),
                "macros_preserved": False,
                "formula_evaluation_performed": False,
            },
        }
    finally:
        formula_book.close()
        value_book.close()


def representatives(inventory: dict[str, Any]) -> list[dict[str, Any]]:
    by_id = {
        str(item.get("file_id")): item
        for item in inventory.get("files", [])
    }
    configured = inventory.get("inventory_representatives") or []
    if configured:
        return [
            item
            for item in (
                by_id.get(str(record.get("representative_file_id")))
                for record in configured
            )
            if item and str(item.get("extension")).casefold() == ".xlsx"
        ]

    seen: set[str] = set()
    result: list[dict[str, Any]] = []
    for item in inventory.get("files", []):
        if str(item.get("extension")).casefold() != ".xlsx":
            continue
        key = str(item.get("sha256") or item.get("relative_path"))
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def build_xlsx_surface_audit(
    input_root: str | Path,
    inventory: dict[str, Any],
) -> dict[str, Any]:
    root = Path(input_root).expanduser().resolve(strict=False)
    if not root.is_dir():
        return {
            "module_id": MODULE_ID,
            "status": "FAIL_CLOSED",
            "hard_block_hits": ["input_root_missing"],
            "canonical_event_count": CANONICAL_EVENT_COUNT,
            "active_match_evidence_pass": False,
            "production_release": False,
            "claim_ceiling": CLAIM_CEILING,
        }

    files: list[dict[str, Any]] = []
    for item in representatives(inventory):
        result = inspect_xlsx_file(
            root / str(item.get("relative_path")),
            str(item.get("source_role") or "UNKNOWN"),
        )
        result.update(
            {
                "file_id": item.get("file_id"),
                "relative_path": item.get("relative_path"),
                "sha256": item.get("sha256"),
                "inventory_sheet_names": item.get("sheet_names") or [],
                "inventory_sheet_states": item.get("sheet_states") or {},
            }
        )
        files.append(result)

    hard_blocks = sorted(
        {
            block
            for result in files
            for block in result.get("hard_block_hits", [])
        }
    )
    if not files:
        status = "FAIL_CLOSED"
        hard_blocks = ["xlsx_surface_missing"]
    elif hard_blocks or any(
        result.get("status") == "FAIL_CLOSED" for result in files
    ):
        status = "FAIL_CLOSED"
    elif any(
        result.get("status") == "REVIEW_REQUIRED" for result in files
    ):
        status = "REVIEW_REQUIRED"
    else:
        status = "PASS"

    return {
        "module_id": MODULE_ID,
        "status": status,
        "input_root": str(root),
        "xlsx_file_count": len(files),
        "files": files,
        "hard_block_hits": hard_blocks,
        "canonical_event_count": CANONICAL_EVENT_COUNT,
        "active_match_evidence_pass": False,
        "production_release": False,
        "claim_ceiling": CLAIM_CEILING,
        "analyst_evidence": {
            "visible_xlsx_surfaces": len(files),
            "safe_statement": (
                "Visible XLSX aggregate surfaces were profiled; aggregate definitions, "
                "identity truth and event truth remain unresolved."
            ),
        },
    }


def validate_out(out_dir: str | Path) -> Path:
    path = Path(out_dir).expanduser().resolve(strict=False)
    if "HPFA" in path.parts and path.name != "HPFA":
        raise ValueError("nested_phone_output_directory_rejected")
    return path


def is_active(path: Path) -> bool:
    return path.as_posix().rstrip("/").endswith(
        "runtime/active_single_match/current"
    )


def write_outputs(
    input_root: str | Path,
    inventory_path: str | Path,
    out_dir: str | Path,
) -> dict[str, Any]:
    output_root = validate_out(out_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    payload = build_xlsx_surface_audit(
        input_root,
        json.loads(Path(inventory_path).read_text(encoding="utf-8")),
    )
    payload["active_match_evidence_pass"] = (
        payload.get("status") != "FAIL_CLOSED"
        and is_active(Path(input_root).resolve(strict=False))
    )
    paths = {key: output_root / value for key, value in OUT.items()}
    payload["outputs"] = {
        key: str(value) for key, value in paths.items()
    }
    paths["main"].write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    paths["summary"].write_text(
        render_summary(payload), encoding="utf-8"
    )
    paths["analyst"].write_text(
        render_analyst(payload), encoding="utf-8"
    )
    return payload


def render_summary(payload: dict[str, Any]) -> str:
    return "\n".join(
        [
            "HPFA XLSX SURFACE READER LITE V1",
            f"status={payload.get('status')}",
            f"xlsx_file_count={payload.get('xlsx_file_count')}",
            f"hard_block_hits={payload.get('hard_block_hits')}",
            (
                "active_match_evidence_pass="
                f"{payload.get('active_match_evidence_pass')}"
            ),
            "canonical_event_count=UNKNOWN",
            "production_release=false",
            "claim_ceiling=XLSX_SURFACE_AUDIT_ONLY",
            "",
        ]
    )


def render_analyst(payload: dict[str, Any]) -> str:
    lines = [
        "HPFA XLSX SURFACE ANALYST AUDIT LITE V1",
        f"status={payload.get('status')}",
        f"visible_xlsx_surfaces={payload.get('xlsx_file_count')}",
    ]
    for result in payload.get("files", []):
        lines += [
            "",
            f"file={result.get('relative_path')}",
            f"source_role={result.get('source_role')}",
            f"status={result.get('status')}",
            f"sheet_count={result.get('sheet_count')}",
            f"hard_block_hits={result.get('hard_block_hits')}",
        ]
        for sheet in result.get("sheets", []):
            metric_labels = [
                item.get("raw_metric_label")
                for item in sheet.get("metric_inventory", [])[:20]
            ]
            lines += [
                f"sheet={sheet.get('sheet_name')}",
                f"sheet_state={sheet.get('sheet_state')}",
                f"surface_rows={sheet.get('surface_row_count')}",
                f"columns={sheet.get('visible_column_count')}",
                f"identity_binding={sheet.get('identity_binding')}",
                (
                    "metric_label_count="
                    f"{len(sheet.get('metric_inventory', []))}"
                ),
                f"metric_labels_sample={metric_labels}",
                f"formula_audit={sheet.get('formula_audit')}",
                f"parse_warnings={sheet.get('parse_warnings')}",
            ]
    return "\n".join(
        lines
        + [
            "",
            "canonical_event_count=UNKNOWN",
            "production_release=false",
            (
                "safe_statement=visible XLSX surface contains aggregate row-level "
                "evidence; aggregate definition truth and event truth remain unresolved."
            ),
            "",
        ]
    )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-root", required=True)
    parser.add_argument("--inventory", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()
    payload = write_outputs(args.input_root, args.inventory, args.out)
    print(
        json.dumps(
            {
                key: payload.get(key)
                for key in (
                    "status",
                    "xlsx_file_count",
                    "hard_block_hits",
                    "active_match_evidence_pass",
                    "canonical_event_count",
                    "production_release",
                )
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 2 if payload.get("status") == "FAIL_CLOSED" else 0


if __name__ == "__main__":
    raise SystemExit(main())
