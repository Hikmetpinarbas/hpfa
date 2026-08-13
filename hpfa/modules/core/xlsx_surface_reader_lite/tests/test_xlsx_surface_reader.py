from __future__ import annotations

import importlib.util
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

SOURCE = (
    Path(__file__).resolve().parents[1]
    / "src"
    / "xlsx_surface_reader.py"
)
SPEC = importlib.util.spec_from_file_location("xlsx_surface_reader_test_module", SOURCE)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
SPEC.loader.exec_module(MODULE)


def make_workbook(path: Path, *, hidden_sheet: bool = False, formula: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Main statistics"
    ws.append(["Player", "Team", "Minutes played", "Passes", "Passes accurate, %"])
    ws.append(["Alpha", "Side A", 90, 50, 80])
    ws.append(["Beta", "Side B", 75, 41, 78])
    if formula:
        ws["D4"] = "=SUM(D2:D3)"
        ws["A4"] = "Total"
    if hidden_sheet:
        hidden = wb.create_sheet("Metadata")
        hidden.sheet_state = "hidden"
        hidden.append(["Key", "Value"])
        hidden.append(["provider", "candidate"])
    wb.save(path)


def inventory_for(path: Path, *, duplicate: bool = False) -> dict:
    item = {
        "file_id": "file_a",
        "relative_path": path.name,
        "extension": ".xlsx",
        "sha256": "same_sha",
        "source_role": "PLAYER_SURFACE_CANDIDATE",
        "sheet_names": ["Main statistics"],
        "sheet_states": {"Main statistics": "visible"},
    }
    files = [item]
    if duplicate:
        files.append(item | {"file_id": "file_b", "relative_path": f"copy/{path.name}"})
    return {"files": files}


def test_visible_sheet_profile_and_identity_candidate_only(tmp_path: Path) -> None:
    path = tmp_path / "players.xlsx"
    make_workbook(path)
    payload = MODULE.build_xlsx_surface_audit(tmp_path, inventory_for(path))
    assert payload["status"] == "PASS"
    assert payload["xlsx_file_count"] == 1
    sheet = payload["files"][0]["sheets"][0]
    assert sheet["surface_row_count"] == 2
    assert sheet["visible_column_count"] == 5
    assert sheet["identity_binding"]["player"]["binding_status"] == "CANDIDATE_ONLY"
    assert sheet["identity_binding"]["player"]["validated_identity"] is False
    assert sheet["identity_binding"]["team"]["validated_identity"] is False
    assert payload["canonical_event_count"] == "UNKNOWN"
    assert payload["production_release"] is False


def test_all_sheet_states_are_audited(tmp_path: Path) -> None:
    path = tmp_path / "players.xlsx"
    make_workbook(path, hidden_sheet=True)
    payload = MODULE.build_xlsx_surface_audit(tmp_path, inventory_for(path))
    assert payload["status"] == "REVIEW_REQUIRED"
    file_audit = payload["files"][0]
    assert file_audit["sheet_count"] == 2
    assert file_audit["visible_sheet_count"] == 1
    assert file_audit["hidden_sheet_count"] == 1
    hidden = next(sheet for sheet in file_audit["sheets"] if sheet["sheet_name"] == "Metadata")
    assert "hidden_sheet_not_admitted_to_visible_surface" in hidden["parse_warnings"]


def test_formula_audit_does_not_evaluate_formula(tmp_path: Path) -> None:
    path = tmp_path / "players.xlsx"
    make_workbook(path, formula=True)
    payload = MODULE.build_xlsx_surface_audit(tmp_path, inventory_for(path))
    sheet = payload["files"][0]["sheets"][0]
    assert sheet["formula_audit"]["formula_cell_count"] == 1
    assert sheet["formula_audit"]["formula_evaluation_performed"] is False
    assert sheet["formula_audit"]["formula_without_cached_value_count"] == 1
    assert payload["status"] == "REVIEW_REQUIRED"


def test_header_discovery_skips_title_row(tmp_path: Path) -> None:
    path = tmp_path / "players.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Main statistics"
    ws.append(["Match report"])
    ws.append([])
    ws.append(["Player", "Team", "Minutes played"])
    ws.append(["Alpha", "Side A", 90])
    wb.save(path)
    payload = MODULE.build_xlsx_surface_audit(tmp_path, inventory_for(path))
    sheet = payload["files"][0]["sheets"][0]
    assert sheet["header_row_index"] == 3
    assert sheet["surface_row_count"] == 1


def test_duplicate_inventory_reflection_is_read_once(tmp_path: Path) -> None:
    path = tmp_path / "players.xlsx"
    make_workbook(path)
    copy_dir = tmp_path / "copy"
    copy_dir.mkdir()
    (copy_dir / path.name).write_bytes(path.read_bytes())
    payload = MODULE.build_xlsx_surface_audit(tmp_path, inventory_for(path, duplicate=True))
    assert payload["xlsx_file_count"] == 1


def test_malformed_xlsx_fails_closed(tmp_path: Path) -> None:
    path = tmp_path / "broken.xlsx"
    path.write_text("not an xlsx", encoding="utf-8")
    payload = MODULE.build_xlsx_surface_audit(tmp_path, inventory_for(path))
    assert payload["status"] == "FAIL_CLOSED"
    assert "malformed_xlsx_container" in payload["hard_block_hits"]


def test_nested_phone_output_directory_rejected() -> None:
    with pytest.raises(ValueError, match="nested_phone_output_directory_rejected"):
        MODULE.validate_out("/sdcard/Download/HPFA/xlsx-run")


def test_write_outputs_and_active_match_evidence(tmp_path: Path) -> None:
    active = tmp_path / "runtime" / "active_single_match" / "current"
    active.mkdir(parents=True)
    path = active / "players.xlsx"
    make_workbook(path)
    inventory = active / "inventory.json"
    inventory.write_text(json.dumps(inventory_for(path)), encoding="utf-8")
    out = tmp_path / "out"
    payload = MODULE.write_outputs(active, inventory, out)
    assert payload["active_match_evidence_pass"] is True
    assert (out / "xlsx_surface_audit_lite_v1.json").is_file()
    assert (out / "xlsx_surface_audit_lite_v1.txt").is_file()
    assert (out / "xlsx_surface_analyst_audit_lite_v1.txt").is_file()


def test_no_sample_match_identity_leak() -> None:
    text = SOURCE.read_text(encoding="utf-8").casefold()
    forbidden = [
        "australia",
        "turkey",
        "galatasaray",
        "fenerbahce",
        "13.06.2026",
        "surface_row_count=31",
        "surface_row_count=3",
    ]
    assert not any(token in text for token in forbidden)


def test_surface_cell_budget_fails_closed_before_iteration() -> None:
    class OversizedSheet:
        title = "Oversized"
        sheet_state = "visible"
        max_row = 2001
        max_column = 1000

    result = MODULE.inspect_sheet(
        OversizedSheet(),
        OversizedSheet(),
        "PLAYER_SURFACE_CANDIDATE",
    )
    assert result["status"] == "FAIL_CLOSED"
    assert result["hard_block_hits"] == ["xlsx_surface_cell_budget_exceeded"]
