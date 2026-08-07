from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import math
import zipfile
from pathlib import Path
from typing import Any

MODULE_ID = "xlsx_entity_metric_row_projection_lite_v1"
CANONICAL_EVENT_COUNT = "UNKNOWN"
CLAIM_CEILING = "XLSX_ROW_ALIGNED_ENTITY_METRIC_SURFACE_ONLY"
ALLOWED_PHONE_OUTPUTS = {
    "/sdcard/Download/HPFA",
    "/storage/emulated/0/Download/HPFA",
}
OUT = {
    "main": "xlsx_entity_metric_row_projection_lite_v1.json",
    "summary": "xlsx_entity_metric_row_projection_lite_v1.txt",
    "analyst": "xlsx_entity_metric_row_projection_analyst_audit_v1.txt",
}
IDENTITY_OUTPUT_KEYS = {
    "player": "player_raw_candidate",
    "team": "team_raw_candidate",
    "position": "position_raw_candidate",
    "minutes": "minutes_raw_candidate",
    "shirt_number": "shirt_number_raw_candidate",
}


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def jsonable(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, float) and not math.isfinite(value):
        return str(value)
    return value


def value_kind(value: Any) -> str:
    if is_blank(value):
        return "blank"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return "date_or_time"
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return "number"
    if isinstance(value, str):
        return "string"
    return type(value).__name__.casefold()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def stable_projection_id(source_sha256: str, sheet_name: str, row_number: int) -> str:
    raw = f"{source_sha256}|{sheet_name}|{row_number}".encode("utf-8")
    return "xrp_" + hashlib.sha256(raw).hexdigest()[:24]


def validate_out(out: str | Path) -> Path:
    raw = str(out)
    if raw.startswith("/sdcard/Download/HPFA/") or raw.startswith(
        "/storage/emulated/0/Download/HPFA/"
    ):
        raise ValueError("nested_phone_output_directory_rejected")
    path = Path(out).expanduser()
    if raw.startswith("/sdcard/") or raw.startswith("/storage/emulated/0/"):
        if raw not in ALLOWED_PHONE_OUTPUTS:
            raise ValueError(f"phone_output_directory_not_allowed:{raw}")
    return path


def _active_match_path(path: Path) -> bool:
    parts = path.resolve(strict=False).parts
    return len(parts) >= 3 and tuple(parts[-3:]) == (
        "runtime",
        "active_single_match",
        "current",
    )


def _inventory_by_id(inventory: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        str(item.get("file_id")): item
        for item in inventory.get("files", [])
        if item.get("file_id") is not None
    }


def _audit_profiles(sheet_audit: dict[str, Any]) -> list[dict[str, Any]]:
    profiles = sheet_audit.get("column_profiles") or []
    expected = int(sheet_audit.get("visible_column_count") or 0)
    if len(profiles) != expected:
        raise ValueError("xlsx_audit_column_profile_count_mismatch")
    return profiles


def _metric_key(profile: dict[str, Any]) -> str:
    return str(profile.get("normalized_column") or "")


def _duplicate_metric_keys(profiles: list[dict[str, Any]]) -> list[str]:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for profile in profiles:
        if profile.get("identity_role_candidate"):
            continue
        key = _metric_key(profile)
        if key in seen:
            duplicates.add(key)
        seen.add(key)
    return sorted(duplicates)


def _validate_file_binding(
    input_root: Path,
    inventory_item: dict[str, Any],
    file_audit: dict[str, Any],
) -> tuple[Path, list[str]]:
    blocks: list[str] = []
    relative = str(file_audit.get("relative_path") or "")
    if not relative:
        blocks.append("xlsx_audit_relative_path_missing")
        return input_root, blocks
    if str(inventory_item.get("relative_path") or "") != relative:
        blocks.append("xlsx_inventory_audit_relative_path_mismatch")
    if str(inventory_item.get("extension") or "").casefold() != ".xlsx":
        blocks.append("inventory_non_xlsx_file_rejected")
    inventory_sha = str(inventory_item.get("sha256") or "")
    audit_sha = str(file_audit.get("sha256") or "")
    if not inventory_sha or inventory_sha != audit_sha:
        blocks.append("xlsx_inventory_audit_sha256_mismatch")
    source_role_inventory = str(inventory_item.get("source_role") or "UNKNOWN")
    source_role_audit = str(file_audit.get("source_role") or "UNKNOWN")
    if source_role_audit not in {"UNKNOWN", source_role_inventory}:
        blocks.append("xlsx_inventory_audit_source_role_mismatch")
    path = input_root / relative
    if not path.is_file():
        blocks.append("xlsx_source_file_missing")
        return path, blocks
    if not zipfile.is_zipfile(path):
        blocks.append("malformed_xlsx_container")
        return path, blocks
    actual_sha = sha256_file(path)
    if actual_sha != inventory_sha:
        blocks.append("xlsx_source_sha256_mismatch")
    return path, blocks


def _project_visible_sheet(
    formula_ws: Any,
    value_ws: Any,
    sheet_audit: dict[str, Any],
    file_meta: dict[str, Any],
    match_surface_binding_id: str | None,
) -> dict[str, Any]:
    hard_blocks: list[str] = []
    review_hits: list[str] = []
    if str(getattr(formula_ws, "sheet_state", "visible")) != "visible":
        hard_blocks.append("audited_visible_sheet_is_not_visible_at_projection_time")
    header_row = int(sheet_audit.get("header_row_index") or 0)
    if header_row <= 0:
        hard_blocks.append("xlsx_audit_header_row_index_missing")
    expected_columns = list(sheet_audit.get("raw_columns") or [])
    profiles = _audit_profiles(sheet_audit)
    duplicates = _duplicate_metric_keys(profiles)
    if duplicates:
        review_hits.extend(f"duplicate_normalized_metric_column:{key}" for key in duplicates)

    if hard_blocks:
        return {
            "sheet_name": formula_ws.title,
            "status": "FAIL_CLOSED",
            "rows": [],
            "hard_block_hits": sorted(set(hard_blocks)),
            "review_hits": sorted(set(review_hits)),
        }

    actual_columns = [
        "" if formula_ws.cell(row=header_row, column=index + 1).value is None
        else str(formula_ws.cell(row=header_row, column=index + 1).value).strip()
        for index in range(len(profiles))
    ]
    if actual_columns != expected_columns:
        return {
            "sheet_name": formula_ws.title,
            "status": "FAIL_CLOSED",
            "rows": [],
            "hard_block_hits": ["xlsx_header_binding_mismatch"],
            "review_hits": sorted(set(review_hits)),
        }

    if duplicates:
        return {
            "sheet_name": formula_ws.title,
            "status": "REVIEW_REQUIRED",
            "rows": [],
            "hard_block_hits": [],
            "review_hits": sorted(set(review_hits)),
            "source_row_count": max(0, int(value_ws.max_row or 0) - header_row),
            "projected_row_count": 0,
        }

    rows: list[dict[str, Any]] = []
    max_row = max(int(formula_ws.max_row or 0), int(value_ws.max_row or 0))
    for row_number in range(header_row + 1, max_row + 1):
        cached_values = [
            value_ws.cell(row=row_number, column=index + 1).value
            for index in range(len(profiles))
        ]
        formulas = [
            formula_ws.cell(row=row_number, column=index + 1)
            for index in range(len(profiles))
        ]
        if not any(
            not is_blank(cached)
            or (cell.data_type == "f" and cell.value is not None)
            for cached, cell in zip(cached_values, formulas)
        ):
            continue

        identity = {key: None for key in IDENTITY_OUTPUT_KEYS.values()}
        metrics: dict[str, dict[str, Any]] = {}
        row_review: list[str] = []
        for index, profile in enumerate(profiles):
            cached = cached_values[index]
            formula_cell = formulas[index]
            formula_present = formula_cell.data_type == "f"
            role = profile.get("identity_role_candidate")
            if role in IDENTITY_OUTPUT_KEYS:
                identity[IDENTITY_OUTPUT_KEYS[str(role)]] = jsonable(cached)
                continue

            key = _metric_key(profile)
            cache_missing = formula_present and is_blank(cached)
            admitted = not cache_missing
            if cache_missing:
                row_review.append(f"formula_without_cached_value:{key}")
            metrics[key] = {
                "raw_metric_label": str(profile.get("raw_column") or ""),
                "raw_value": jsonable(cached) if admitted else None,
                "value_kind": value_kind(cached) if admitted else "blank",
                "number_format": str(
                    value_ws.cell(row=row_number, column=index + 1).number_format
                    or formula_cell.number_format
                    or ""
                ),
                "percent_header_candidate": bool(profile.get("percent_header_candidate")),
                "formula_present": formula_present,
                "cached_value_used": bool(formula_present and admitted),
                "value_admitted": admitted,
                "value_status": (
                    "NOT_ADMITTED_FORMULA_CACHE_MISSING"
                    if cache_missing
                    else ("MISSING" if is_blank(cached) else "OBSERVED")
                ),
            }

        rows.append(
            {
                "row_projection_id": stable_projection_id(
                    str(file_meta["source_sha256"]), formula_ws.title, row_number
                ),
                "file_id": file_meta["file_id"],
                "relative_path": file_meta["relative_path"],
                "source_sha256": file_meta["source_sha256"],
                "source_role": file_meta["source_role"],
                "sheet_name": formula_ws.title,
                "sheet_state": "visible",
                "header_row_index": header_row,
                "source_row_number": row_number,
                "match_surface_binding_id": match_surface_binding_id,
                "identity_candidates": identity,
                "metric_values": metrics,
                "row_surface_claim_ceiling": CLAIM_CEILING,
                "validated_identity": False,
                "canonical_event_count": CANONICAL_EVENT_COUNT,
                "production_release": False,
                "review_hits": sorted(set(row_review)),
            }
        )

    status = "REVIEW_REQUIRED" if any(row["review_hits"] for row in rows) else "PASS"
    return {
        "sheet_name": formula_ws.title,
        "status": status,
        "rows": rows,
        "source_row_count": len(rows),
        "projected_row_count": len(rows),
        "hard_block_hits": [],
        "review_hits": sorted(
            {hit for row in rows for hit in row.get("review_hits", [])}
        ),
    }


def build_projection(
    input_root: str | Path,
    inventory: dict[str, Any],
    xlsx_audit: dict[str, Any],
    *,
    match_surface_binding_id: str | None = None,
) -> dict[str, Any]:
    root = Path(input_root).expanduser().resolve(strict=False)
    base = {
        "module_id": MODULE_ID,
        "claim_ceiling": CLAIM_CEILING,
        "match_surface_binding_id": match_surface_binding_id,
        "canonical_event_count": CANONICAL_EVENT_COUNT,
        "production_release": False,
        "validated_identity": False,
        "aggregate_definition_truth": False,
        "metric_truth": False,
        "claim_allowed": False,
    }
    if not root.is_dir():
        return base | {
            "status": "FAIL_CLOSED",
            "hard_block_hits": ["input_root_missing"],
            "review_hits": [],
            "files": [],
            "row_projection_count": 0,
        }
    if str(xlsx_audit.get("module_id") or "") != "xlsx_surface_reader_lite_v1":
        return base | {
            "status": "FAIL_CLOSED",
            "hard_block_hits": ["xlsx_audit_module_id_mismatch"],
            "review_hits": [],
            "files": [],
            "row_projection_count": 0,
        }
    if xlsx_audit.get("status") == "FAIL_CLOSED":
        return base | {
            "status": "FAIL_CLOSED",
            "hard_block_hits": ["xlsx_surface_audit_fail_closed"],
            "review_hits": [],
            "files": [],
            "row_projection_count": 0,
        }

    by_id = _inventory_by_id(inventory)
    files: list[dict[str, Any]] = []
    global_blocks: list[str] = []
    global_reviews: list[str] = []

    for file_audit in xlsx_audit.get("files", []):
        file_id = str(file_audit.get("file_id") or "")
        item = by_id.get(file_id)
        if item is None:
            global_blocks.append(f"xlsx_inventory_file_id_missing:{file_id}")
            continue
        path, blocks = _validate_file_binding(root, item, file_audit)
        source_role = str(item.get("source_role") or "UNKNOWN")
        file_result: dict[str, Any] = {
            "file_id": file_id,
            "relative_path": str(file_audit.get("relative_path") or ""),
            "source_sha256": str(file_audit.get("sha256") or ""),
            "source_role": source_role,
            "status": "FAIL_CLOSED" if blocks else "PASS",
            "sheets": [],
            "hard_block_hits": sorted(set(blocks)),
            "review_hits": [],
        }
        if blocks:
            files.append(file_result)
            global_blocks.extend(blocks)
            continue

        try:
            from openpyxl import load_workbook
            formula_book = load_workbook(path, read_only=False, data_only=False, keep_links=False)
            value_book = load_workbook(path, read_only=False, data_only=True, keep_links=False)
        except ModuleNotFoundError:
            block = "openpyxl_dependency_missing"
            file_result["status"] = "FAIL_CLOSED"
            file_result["hard_block_hits"] = [block]
            files.append(file_result)
            global_blocks.append(block)
            continue
        except Exception:
            block = "malformed_or_unreadable_xlsx"
            file_result["status"] = "FAIL_CLOSED"
            file_result["hard_block_hits"] = [block]
            files.append(file_result)
            global_blocks.append(block)
            continue

        try:
            audit_sheets = file_audit.get("sheets") or []
            for sheet_audit in audit_sheets:
                name = str(sheet_audit.get("sheet_name") or "")
                audited_state = str(sheet_audit.get("sheet_state") or "visible")
                if audited_state != "visible":
                    file_result["review_hits"].append(f"hidden_sheet_not_projected:{name}")
                    continue
                if name not in formula_book.sheetnames or name not in value_book.sheetnames:
                    block = f"audited_sheet_missing:{name}"
                    file_result["hard_block_hits"].append(block)
                    global_blocks.append(block)
                    continue
                try:
                    sheet_result = _project_visible_sheet(
                        formula_book[name],
                        value_book[name],
                        sheet_audit,
                        {
                            "file_id": file_id,
                            "relative_path": file_result["relative_path"],
                            "source_sha256": file_result["source_sha256"],
                            "source_role": source_role,
                        },
                        match_surface_binding_id,
                    )
                except ValueError as exc:
                    sheet_result = {
                        "sheet_name": name,
                        "status": "FAIL_CLOSED",
                        "rows": [],
                        "hard_block_hits": [str(exc)],
                        "review_hits": [],
                    }
                file_result["sheets"].append(sheet_result)
                file_result["hard_block_hits"].extend(sheet_result.get("hard_block_hits", []))
                file_result["review_hits"].extend(sheet_result.get("review_hits", []))
                global_blocks.extend(sheet_result.get("hard_block_hits", []))
                global_reviews.extend(sheet_result.get("review_hits", []))
        finally:
            formula_book.close()
            value_book.close()

        file_result["hard_block_hits"] = sorted(set(file_result["hard_block_hits"]))
        file_result["review_hits"] = sorted(set(file_result["review_hits"]))
        if file_result["hard_block_hits"]:
            file_result["status"] = "FAIL_CLOSED"
        elif file_result["review_hits"] or any(
            sheet.get("status") == "REVIEW_REQUIRED" for sheet in file_result["sheets"]
        ):
            file_result["status"] = "REVIEW_REQUIRED"
        else:
            file_result["status"] = "PASS"
        files.append(file_result)
        global_reviews.extend(file_result["review_hits"])

    row_count = sum(
        len(sheet.get("rows", []))
        for file_row in files
        for sheet in file_row.get("sheets", [])
    )
    if not files:
        global_blocks.append("xlsx_projection_source_missing")
    if global_blocks:
        status = "FAIL_CLOSED"
    elif global_reviews or any(file_row.get("status") == "REVIEW_REQUIRED" for file_row in files):
        status = "REVIEW_REQUIRED"
    else:
        status = "PASS"
    return base | {
        "status": status,
        "input_root": str(root),
        "xlsx_file_count": len(files),
        "row_projection_count": row_count,
        "files": files,
        "hard_block_hits": sorted(set(global_blocks)),
        "review_hits": sorted(set(global_reviews)),
        "does_not_measure": [
            "canonical_event_truth",
            "validated_player_identity",
            "validated_team_identity",
            "aggregate_definition_truth",
            "metric_truth",
            "comparison_truth",
            "sequence_truth",
            "phase_truth",
            "tactical_truth",
        ],
    }


def _load_json(path: str | Path) -> dict[str, Any]:
    with Path(path).open(encoding="utf-8") as handle:
        return json.load(handle)


def _runtime_status(payload_status: str, active: bool) -> str:
    if not active:
        return "NOT_EVALUATED"
    suffix = {
        "PASS": "PASS",
        "REVIEW_REQUIRED": "REVIEW_REQUIRED",
        "FAIL_CLOSED": "FAIL_CLOSED",
    }.get(payload_status, "REVIEW_REQUIRED")
    return f"ACTIVE_MATCH_EXECUTION_COMPLETED_{suffix}"


def write_outputs(
    input_root: str | Path,
    inventory_path: str | Path,
    xlsx_audit_path: str | Path,
    out: str | Path,
    *,
    runtime_authority: str | Path | None = None,
    active_match_execution: bool = False,
    match_surface_binding_id: str | None = None,
) -> dict[str, Any]:
    out_path = validate_out(out)
    root = Path(input_root).expanduser().resolve(strict=False)
    if active_match_execution:
        authority = Path(runtime_authority or input_root).expanduser().resolve(strict=False)
        if root != authority or not _active_match_path(authority):
            raise ValueError("active_match_runtime_authority_mismatch")
    payload = build_projection(
        root,
        _load_json(inventory_path),
        _load_json(xlsx_audit_path),
        match_surface_binding_id=match_surface_binding_id,
    )
    payload["runtime_evidence_status"] = _runtime_status(
        str(payload.get("status")), active_match_execution
    )
    payload["active_match_evidence_pass"] = bool(
        active_match_execution and payload.get("status") != "FAIL_CLOSED"
    )
    payload["runtime_authority"] = (
        str(Path(runtime_authority or input_root).expanduser().resolve(strict=False))
        if active_match_execution
        else None
    )
    out_path.mkdir(parents=True, exist_ok=True)
    (out_path / OUT["main"]).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    summary = "\n".join(
        [
            "HPFA XLSX ENTITY-METRIC ROW PROJECTION LITE V1",
            f"status={payload.get('status')}",
            f"runtime_evidence_status={payload.get('runtime_evidence_status')}",
            f"xlsx_file_count={payload.get('xlsx_file_count')}",
            f"row_projection_count={payload.get('row_projection_count')}",
            f"hard_block_hits={payload.get('hard_block_hits')}",
            f"review_hits={payload.get('review_hits')}",
            f"canonical_event_count={CANONICAL_EVENT_COUNT}",
            "production_release=false",
        ]
    ) + "\n"
    (out_path / OUT["summary"]).write_text(summary, encoding="utf-8")
    analyst_lines = [
        "HPFA XLSX ENTITY-METRIC ROW PROJECTION ANALYST AUDIT",
        "Row-level evidence only: identity candidates and metric cells were observed on the same visible XLSX row.",
        "This does not validate player/team identity, aggregate definition, metric truth, or comparison truth.",
        f"status={payload.get('status')}",
        f"row_projection_count={payload.get('row_projection_count')}",
    ]
    for file_row in payload.get("files", []):
        for sheet in file_row.get("sheets", []):
            analyst_lines.append(
                "file={relative} sheet={sheet} status={status} rows={rows}".format(
                    relative=file_row.get("relative_path"),
                    sheet=sheet.get("sheet_name"),
                    status=sheet.get("status"),
                    rows=len(sheet.get("rows", [])),
                )
            )
    analyst_lines.extend(["canonical_event_count=UNKNOWN", "production_release=false"])
    (out_path / OUT["analyst"]).write_text(
        "\n".join(analyst_lines) + "\n", encoding="utf-8"
    )
    return payload


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-root", required=True)
    parser.add_argument("--inventory", required=True)
    parser.add_argument("--xlsx-audit", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--runtime-authority")
    parser.add_argument("--active-match-execution", action="store_true")
    parser.add_argument("--match-surface-binding-id")
    args = parser.parse_args(argv)
    try:
        payload = write_outputs(
            args.input_root,
            args.inventory,
            args.xlsx_audit,
            args.out,
            runtime_authority=args.runtime_authority,
            active_match_execution=args.active_match_execution,
            match_surface_binding_id=args.match_surface_binding_id,
        )
    except ValueError as exc:
        print(
            json.dumps(
                {
                    "module_id": MODULE_ID,
                    "status": "FAIL_CLOSED",
                    "hard_block_hits": [str(exc)],
                    "active_match_evidence_pass": False,
                    "canonical_event_count": CANONICAL_EVENT_COUNT,
                    "production_release": False,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 2
    print(
        json.dumps(
            {
                "module_id": MODULE_ID,
                "status": payload.get("status"),
                "runtime_evidence_status": payload.get("runtime_evidence_status"),
                "row_projection_count": payload.get("row_projection_count"),
                "canonical_event_count": CANONICAL_EVENT_COUNT,
                "production_release": False,
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 2 if payload.get("status") == "FAIL_CLOSED" else 0


if __name__ == "__main__":
    raise SystemExit(main())
